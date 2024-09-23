# gui/overview_tab.py
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QPushButton, QTableWidget, QTableWidgetItem,
    QHBoxLayout, QMessageBox, QFileDialog
)
from PyQt5.QtCore import Qt
from gui.add_project_dialog import AddProjectDialog
from utils import sanitize_filename
import sys
import os
import subprocess

from utils import get_template_dir

template_dir = get_template_dir()

from pdf_converter import PDFConverter
from PyQt5.QtCore import QThread
import tempfile
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

class OverviewTab(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        # Buttons Layout
        self.buttons_layout = QHBoxLayout()
        
        # Add New Project Button
        self.add_project_btn = QPushButton("Add New Project")
        self.add_project_btn.clicked.connect(self.open_add_project_dialog)
        self.buttons_layout.addWidget(self.add_project_btn)

        # Send to Printer Button
        self.send_printer_btn = QPushButton("Send to Printer")
        self.send_printer_btn.clicked.connect(self.send_to_printer)
        self.buttons_layout.addWidget(self.send_printer_btn)

        # Open PDF Button
        self.open_pdf_btn = QPushButton("Open PDF")
        self.open_pdf_btn.clicked.connect(self.open_pdf)
        self.buttons_layout.addWidget(self.open_pdf_btn)

        self.layout.addLayout(self.buttons_layout)

        # Projects Table
        self.table = QTableWidget()
        self.table.setColumnCount(10)  # Updated column count
        self.table.setHorizontalHeaderLabels([
            "Project Name",
            "Project Number",
            "Complex",
            "Start Date",
            "End Date",
            "Status",
            "Worker",
            "Innregulering",
            "Sjekkliste",
            "Move"
        ])
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.layout.addWidget(self.table)

        # Temporary PDF path
        self.pdf_path = os.path.join(tempfile.gettempdir(), "Project_Overview.pdf")

        self.load_projects()

    def load_projects(self):
        self.table.setRowCount(0)
        projects = self.db.load_projects(status="Active")
        for project in projects:
            row_position = self.table.rowCount()
            self.table.insertRow(row_position)

            self.table.setItem(row_position, 0, QTableWidgetItem(project.name))
            self.table.setItem(row_position, 1, QTableWidgetItem(project.number))
            complex_text = "Yes" if project.is_residential_complex else "No"
            self.table.setItem(row_position, 2, QTableWidgetItem(complex_text))
            self.table.setItem(row_position, 3, QTableWidgetItem(self.format_date(project.start_date)))
            self.table.setItem(row_position, 4, QTableWidgetItem(self.format_date(project.end_date) if project.end_date else ""))
            self.table.setItem(row_position, 5, QTableWidgetItem(project.status))
            self.table.setItem(row_position, 6, QTableWidgetItem(project.worker))

            # Innregulering Button
            innregulering_btn = QPushButton("View PDF")
            innregulering_btn.clicked.connect(lambda checked, p=project: self.view_pdf(p, "Innregulering"))
            self.table.setCellWidget(row_position, 7, innregulering_btn)

            # Sjekkliste Button
            sjekkliste_btn = QPushButton("View PDF")
            sjekkliste_btn.clicked.connect(lambda checked, p=project: self.view_pdf(p, "Sjekkliste"))
            self.table.setCellWidget(row_position, 8, sjekkliste_btn)

            # Move to Complete Button
            move_complete_btn = QPushButton("Complete")
            move_complete_btn.setStyleSheet("background-color: yellow")
            move_complete_btn.clicked.connect(lambda checked, p=project: self.move_to_complete(p))
            self.table.setCellWidget(row_position, 9, move_complete_btn)

    def format_date(self, date_str):
        try:
            return datetime.strptime(date_str, "%Y-%m-%d").strftime("%d-%m-%Y")
        except:
            return date_str

    def open_add_project_dialog(self):
        dialog = AddProjectDialog(self.db)
        if dialog.exec_() == dialog.Accepted:
            self.load_projects()

    def view_pdf(self, project, doc_type):
        folder_name = sanitize_filename(f"{project.name}_{project.number}")
        project_folder = os.path.join("Boligventilasjon - Prosjekter", folder_name)
        pdf_file = os.path.join(project_folder, f"{doc_type}.pdf")

        if not os.path.exists(pdf_file):
            QMessageBox.warning(self, "PDF Not Found", f"The PDF for {doc_type} does not exist.")
            return

        try:
            if sys.platform.startswith('darwin'):
                subprocess.call(('open', pdf_file))
            elif os.name == 'nt':
                os.startfile(pdf_file)
            elif os.name == 'posix':
                subprocess.call(('xdg-open', pdf_file))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open PDF: {str(e)}")

    def move_to_complete(self, project):
        project.status = "Complete"
        self.db.update_project(project)
        self.load_projects()
        QMessageBox.information(self, "Status Updated", f"Project '{project.name}' moved to Complete.")

    def save_pdf(self):
        try:
            # Define Template directory
            script_dir = os.path.dirname(os.path.abspath(__file__))
            template_dir = os.path.join(script_dir, "Template")

            # Check if 'Template' directory exists
            if not os.path.exists(template_dir):
                QMessageBox.critical(self, "Template Missing", f"The 'Template' directory does not exist at:\n{template_dir}\nPlease run 'Setup Template' from the File menu.")
                return

            # Check if required template files exist
            required_files = ["Innregulering.xlsx", "Sjekkliste.xlsx"]
            missing_files = [f for f in required_files if not os.path.exists(os.path.join(template_dir, f))]

            if missing_files:
                QMessageBox.critical(
                    self,
                    "Template Files Missing",
                    f"The following template files are missing in the 'Template' folder:\n" + "\n".join(missing_files) + "\nPlease run 'Setup Template' from the File menu."
                )
                return

            # Create a new Excel workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Project Overview"

            # Write headers
            headers = ["Project Name", "Project Number", "Complex", "Start Date", "End Date", "Status", "Worker"]
            sheet.append(headers)

            # Write data
            for row in range(self.table.rowCount()):
                row_data = []
                for column in range(7):  # Exclude buttons
                    item = self.table.item(row, column)
                    row_data.append(item.text() if item else "")
                sheet.append(row_data)

            # Add current date at the bottom right
            current_date = datetime.now().strftime("%d-%m-%Y")
            last_row = sheet.max_row + 2
            sheet.cell(row=last_row, column=7, value=f"Generated on: {current_date}")
            sheet.cell(row=last_row, column=7).alignment = Alignment(horizontal="right")

            # Save Excel file
            excel_path = os.path.join(tempfile.gettempdir(), "Project_Overview.xlsx")
            workbook.save(excel_path)

            # Convert Excel to PDF
            converter = PDFConverter(excel_path, self.pdf_path)
            converter.run_conversion()

            QMessageBox.information(self, "Success", f"PDF saved successfully at {self.pdf_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save PDF: {str(e)}")

    def open_pdf(self):
        if not os.path.exists(self.pdf_path):
            QMessageBox.warning(self, "PDF Not Found", "No PDF has been saved yet.")
            return

        try:
            if sys.platform.startswith('darwin'):
                subprocess.call(('open', self.pdf_path))
            elif os.name == 'nt':
                os.startfile(self.pdf_path)
            elif os.name == 'posix':
                subprocess.call(('xdg-open', self.pdf_path))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open PDF: {str(e)}")

    def send_to_printer(self):
        try:
            # Save PDF before printing
            self.save_pdf()

            if not os.path.exists(self.pdf_path):
                QMessageBox.warning(self, "PDF Not Found", "PDF could not be created.")
                return

            # Send PDF to printer
            if sys.platform.startswith('darwin'):
                subprocess.call(['lp', self.pdf_path])
            elif os.name == 'nt':
                # Windows doesn't have a direct print command, use ShellExecute
                os.startfile(self.pdf_path, "print")
            elif os.name == 'posix':
                subprocess.call(['lp', self.pdf_path])
            
            QMessageBox.information(self, "Print", "The PDF has been sent to the printer.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to send to printer: {str(e)}")
